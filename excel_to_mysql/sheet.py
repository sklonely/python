import openpyxl

wb = openpyxl.load_workbook('mysql_table.xlsx')
print("sheet Loading OK")


def sheetGetTable(sheetname):  # 從sheet獲取table名稱的list
    sheet1 = wb[sheetname]  # 開啟指定的sheet
    tableName = []
    for cell in list(sheet1.rows)[0]:  # 讀取excel row 1 全部資料(table)
        tableName.append(cell.value)
    return tableName


def sheetGetData(sheetname):  # 從sheet獲取所有資料的list
    sheet1 = wb[sheetname]  # 開啟指定的sheet
    row = range(1, sheet1.max_row)  # 獲取sheet內資料筆數
    data = []
    for x in row:
        data.append([])
        for cell in list(sheet1.rows)[x]:  # 讀取excel row 1 全部資料(table)
            data[x - 1].append(cell.value)
    return data


def sheetNames():
    return wb.sheetnames
